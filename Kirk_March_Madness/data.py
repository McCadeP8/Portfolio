import pandas as pd
import openpyxl
import numpy as np
import streamlit.components.v1 as components
import plotly.graph_objects as go
from functools import reduce
import streamlit as st

import os

def actual_results(Sims):
    cols = [c for c in Sims.columns if c != "Sim"]
    actual = {"Sim": len(Sims) + 1}
    for col in cols:
        unique = Sims[col].dropna().unique()
        actual[col] = unique[0] if len(unique) == 1 else np.nan
    return pd.DataFrame([actual])

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

@st.cache_data
def get_sims_pre():
    file_path = os.path.join(BASE_DIR, "SimsPre.parquet")
    #file_path = ("SimsPre.parquet")
    SimsPre = pd.read_parquet(file_path)
    return SimsPre

@st.cache_data
def get_risk():
    file_path = os.path.join(BASE_DIR, "RiskScore.parquet")
    #file_path = ("RiskScore.parquet")
    Risk = pd.read_parquet(file_path)
    return Risk

@st.cache_data()
def get_picks(Projections):
    file_path = os.path.join(BASE_DIR, "mm 2026 name.xlsm")
    #file_path = ("mm 2026 name.xlsm")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    cells = []
    for r in [4,8,12,16,20,24,28,32,38,42,46,50,54,58,62,66]:
        cells.append(f"F{r}")
    for r in [4,8,12,16,20,24,28,32,38,42,46,50,54,58,62,66]:
        cells.append(f"V{r}")
    for r in [6,14,22,30,40,48,56,64]:
        cells.append(f"H{r}")
    for r in [6,14,22,30,40,48,56,64]:
        cells.append(f"T{r}")
    for r in [10,26,44,60]:
        cells.append(f"J{r}")
    for r in [10,26,44,60]:
        cells.append(f"R{r}")
    cells += ["L17","L51","P17","P51"]
    cells += ["N32","N38","N35"]
    cells += ["L1"]
    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        values = [ws[cell].value for cell in cells]
        rows.append([sheet] + values)
    result = pd.DataFrame(rows, columns=["SheetName"] + cells)
    result = result[(result["N35"].notna()) & (result["N35"] != 54)]
    result = result.drop(columns=["SheetName"])
    result = result.rename(columns={
        "F4": "R64_1",
        "F8": "R64_2",
        "F12": "R64_3",
        "F16": "R64_4",
        "F20": "R64_5",
        "F24": "R64_6",
        "F28": "R64_7",
        "F32": "R64_8",
        "F38": "R64_9",
        "F42": "R64_10",
        "F46": "R64_11",
        "F50": "R64_12",
        "F54": "R64_13",
        "F58": "R64_14",
        "F62": "R64_15",
        "F66": "R64_16",
        "V4": "R64_17",
        "V8": "R64_18",
        "V12": "R64_19",
        "V16": "R64_20",
        "V20": "R64_21",
        "V24": "R64_22",
        "V28": "R64_23",
        "V32": "R64_24",
        "V38": "R64_25",
        "V42": "R64_26",
        "V46": "R64_27",
        "V50": "R64_28",
        "V54": "R64_29",
        "V58": "R64_30",
        "V62": "R64_31",
        "V66": "R64_32",
        "H6": "R32_1",
        "H14": "R32_2",
        "H22": "R32_3",
        "H30": "R32_4",
        "H40": "R32_5",
        "H48": "R32_6",
        "H56": "R32_7",
        "H64": "R32_8",
        "T6": "R32_9",
        "T14": "R32_10",
        "T22": "R32_11",
        "T30": "R32_12",
        "T40": "R32_13",
        "T48": "R32_14",
        "T56": "R32_15",
        "T64": "R32_16",
        "J10": "S16_1",
        "J26": "S16_2",
        "J44": "S16_3",
        "J60": "S16_4",
        "R10": "S16_5",
        "R26": "S16_6",
        "R44": "S16_7",
        "R60": "S16_8",
        "L17": "E8_1",
        "L51": "E8_2",
        "P17": "E8_3",
        "P51": "E8_4",
        "N32": "F4_1",
        "N38": "F4_2",
        "N35": "Champ",
        "L1": "Bracket"})

    team_map = dict(zip(Projections["Team"], Projections["ActualName"]))
    cols = result.columns.difference(["Bracket"])
    result[cols] = result[cols].replace(team_map)
    result["Bracket"] = (
    result["Bracket"]
    + " "
    + (result.groupby("Bracket").cumcount() + 1).astype(str))
    return result

@st.cache_data
def get_projections():
    csv_url = "https://docs.google.com/spreadsheets/d/12f4bu9JRwZ9TDXVw6T2GI0fPgjeKCk1GxrdDFdjHds8/export?format=csv&gid=1837691522"
    df = pd.read_csv(csv_url)
    return df

def calculate_risk_score(projections, picks):
    round_points = {'R64': 1, 'R32': 4, 'S16': 8, 'E8': 16, 'F4': 28, 'Champ': 40}
    round_prob_col = {'R64': 'R32', 'R32': 'S16', 'S16': 'E8', 'E8': 'F4', 'F4': 'Champ', 'Champ': 'Champ'}
    round_cols = {
        'R64':   [f'R64_{i}' for i in range(1, 33)],
        'R32':   [f'R32_{i}' for i in range(1, 17)],
        'S16':   [f'S16_{i}' for i in range(1, 9)],
        'E8':    [f'E8_{i}'  for i in range(1, 5)],
        'F4':    ['F4_1', 'F4_2'],
        'Champ': ['Champ']}
    team_info = projections.set_index('Team')
    results = []
    for _, row in picks.iterrows():
        downside      = 0.0
        total_exp     = 0.0
        champ_exp     = 0.0
        upset_num     = 0.0
        upset_denom   = 0.0
        champ_team = row.get('Champ')
        for round_name, cols in round_cols.items():
            base_pts = round_points[round_name]
            prob_col = round_prob_col[round_name]
            for col in cols:
                team = row[col]
                if pd.isna(team) or team not in team_info.index:
                    continue
                seed     = team_info.loc[team, 'Seed']
                prob     = team_info.loc[team, prob_col]
                pts      = base_pts + seed
                exp_pts  = prob * pts
                downside += (1 - prob) * pts
                total_exp += exp_pts
                if pd.notna(champ_team) and team == champ_team:
                    champ_exp += exp_pts                
                upset_num   += seed * base_pts
                upset_denom += base_pts
        champ_concentration = (champ_exp / total_exp * 100) if total_exp > 0 else 0
        avg_upset_seed      = (upset_num / upset_denom) if upset_denom > 0 else 0
        results.append({
            'downside':            downside,
            'champ_concentration': champ_concentration,
            'avg_upset_seed':      avg_upset_seed})
    df = pd.DataFrame(results)
    def normalize(col):
        mn, mx = col.min(), col.max()
        return (col - mn) / (mx - mn) * 100 if mx > mn else col * 0
    df['downside_score']       = normalize(df['downside'])
    df['concentration_score']  = normalize(df['champ_concentration'])
    df['upset_score']          = normalize(df['avg_upset_seed'])
    df['risk_score'] = (
        df['downside_score']      * 0.50 +
        df['concentration_score'] * 0.25 +
        df['upset_score']         * 0.25)
    df['Bracket']   = picks['Bracket'].values
    df['risk_rank'] = df['risk_score'].rank(ascending=False).astype(int)
    df['downside_rank'] = df['downside_score'].rank(ascending=False).astype(int)
    df['concentration_rank'] = df['concentration_score'].rank(ascending=False).astype(int)
    df['upset_score_rank'] = df['upset_score'].rank(ascending=False).astype(int)
    return df[['Bracket', 'risk_score', 'risk_rank', 'downside_rank', 'concentration_rank', 'upset_score_rank',
               'downside', 'champ_concentration', 'avg_upset_seed']].sort_values('risk_rank')


def get_risk_value(RiskScore, SelectedTeam, SelectedColumn):
    filtered = RiskScore[RiskScore["Bracket"] == SelectedTeam]
    if filtered.empty:
        return None
    return filtered.iloc[0][SelectedColumn]

@st.cache_data
def run_simulations(projections, n_simulations=100000):
    teams = projections['Team'].values
    group_map = projections.set_index('Team')
    def draw_round(prob_col, group_col, n_groups):
        draws = {}
        for g in range(1, n_groups + 1):
            g_teams = projections[projections[group_col] == g]
            probs = g_teams[prob_col].values / g_teams[prob_col].sum()
            draws[g] = np.random.choice(g_teams['Team'].values, size=n_simulations, p=probs)
        return draws
    champ_probs = projections['Champ'].values / projections['Champ'].sum()
    champions = np.random.choice(teams, size=n_simulations, p=champ_probs)
    f4  = draw_round('F4',  'F4Group',  2)
    e8  = draw_round('E8',  'E8Group',  4)
    s16 = draw_round('S16', 'S16Group', 8)
    r32 = draw_round('R32', 'R32Group', 16)
    r64 = draw_round('R64', 'R64Group', 32)
    champ_f4 = group_map.loc[champions, 'F4Group'].values
    for g in [1, 2]:
        f4[g] = np.where(champ_f4 == g, champions, f4[g])
    for g in [1, 2]:
        winner_e8 = group_map.loc[f4[g], 'E8Group'].values
        for eg in [2*g-1, 2*g]:
            e8[eg] = np.where(winner_e8 == eg, f4[g], e8[eg])
    for g in range(1, 5):
        winner_s16 = group_map.loc[e8[g], 'S16Group'].values
        for sg in [2*g-1, 2*g]:
            s16[sg] = np.where(winner_s16 == sg, e8[g], s16[sg])
    for g in range(1, 9):
        winner_r32 = group_map.loc[s16[g], 'R32Group'].values
        for rg in [2*g-1, 2*g]:
            r32[rg] = np.where(winner_r32 == rg, s16[g], r32[rg])
    for g in range(1, 17):
        winner_r64 = group_map.loc[r32[g], 'R64Group'].values
        for rg in [2*g-1, 2*g]:
            r64[rg] = np.where(winner_r64 == rg, r32[g], r64[rg])
    results = pd.DataFrame({'Sim': range(1, n_simulations + 1), 'Champ': champions})
    for g in [1, 2]:          
        results[f'F4_{g}']  = f4[g]
    for g in range(1, 5):   
        results[f'E8_{g}']  = e8[g]
    for g in range(1, 9):     
        results[f'S16_{g}'] = s16[g]
    for g in range(1, 17):    
        results[f'R32_{g}'] = r32[g]
    for g in range(1, 33):    
        results[f'R64_{g}'] = r64[g]
    return results

def score_simulations_by_round(picks, simulations, projections):
    round_points = {'R64': 1, 'R32': 4, 'S16': 8, 'E8': 16, 'F4': 28, 'Champ': 40}
    round_cols = {
        'R64':   [f'R64_{i}' for i in range(1, 33)],
        'R32':   [f'R32_{i}' for i in range(1, 17)],
        'S16':   [f'S16_{i}' for i in range(1, 9)],
        'E8':    [f'E8_{i}'  for i in range(1, 5)],
        'F4':    ['F4_1', 'F4_2'],
        'Champ': ['Champ']}
    seed_map = projections.set_index("Team")["Seed"]
    results = {}
    total_scores = None
    for rnd, cols in round_cols.items():
        sim_arr   = simulations[cols].values
        picks_arr = picks[cols].values
        seed_values = np.vectorize(lambda x: seed_map.get(x, 0))(picks_arr)
        pick_points = seed_values + round_points[rnd]
        match = sim_arr[:, None, :] == picks_arr[None, :, :]        
        scores = (match * pick_points[None, :, :]).sum(axis=2)        
        df = pd.DataFrame(scores, columns=picks["Bracket"].values)
        results[rnd] = df
        if total_scores is None:
            total_scores = scores.copy()        
        else:
            total_scores += scores

    total_df = pd.DataFrame(total_scores, columns=picks["Bracket"].values)
    total_df.insert(0, "Sim", np.arange(1, len(total_df) + 1))
    results["Total"] = total_df

    return (results["R64"], results["R32"], results["S16"], results["E8"], results["F4"], results["Champ"], results["Total"])

def count_simulations_by_round(picks, simulations):
    round_cols = {
        'R64':   [f'R64_{i}' for i in range(1, 33)],
        'R32':   [f'R32_{i}' for i in range(1, 17)],
        'S16':   [f'S16_{i}' for i in range(1, 9)],
        'E8':    [f'E8_{i}'  for i in range(1, 5)],
        'F4':    ['F4_1', 'F4_2'],
        'Champ': ['Champ']}
    results = {}
    total_counts = None
    for rnd, cols in round_cols.items():
        sim_arr   = simulations[cols].values
        picks_arr = picks[cols].values
        match = sim_arr[:, None, :] == picks_arr[None, :, :]
        counts = match.sum(axis=2)
        df = pd.DataFrame(counts, columns=picks["Bracket"].values)
        df.insert(0, "Sim", np.arange(1, len(df) + 1))
        results[rnd] = df
        if total_counts is None:
            total_counts = counts.copy()
        else:
            total_counts += counts
    total_df = pd.DataFrame(total_counts, columns=picks["Bracket"].values)
    total_df.insert(0, "Sim", np.arange(1, len(total_df) + 1))
    results["Total"] = total_df
    return (
        results["R64"],
        results["R32"],
        results["S16"],
        results["E8"],
        results["F4"],
        results["Champ"],
        results["Total"])

def calculate_sim_ranks(scores_total):
    ranks_df = (
        scores_total.drop(columns="Sim")
        .rank(axis=1, ascending=False, method="min")
        .astype(int))
    ranks_df.insert(0, "Sim", np.arange(1, len(ranks_df) + 1))
    return ranks_df

def plot_correct_picks(counts_df, selected_bracket, title, actual=None):
    counts = counts_df[selected_bracket]
    mean_val = counts.mean()
    min_val = counts.min()
    max_val = counts.max()
    bins = np.arange(min_val, max_val + 2)
    values, bins = np.histogram(counts, bins=bins)
    pct = values / values.sum() * 100
    x = bins[:-1]
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=x,
        y=pct,
        marker_color='#009CDE',
        marker_line_width=0))
    fig.add_vline(x=mean_val, line_dash='dash', line_color='red', line_width=1.5)
    fig.add_annotation(
        x=mean_val, y=1.02, yref='paper',
        text="Expected Pre-Tournament",
        showarrow=False,
        xanchor='left', yanchor='top',
        font=dict(color='red', size=12))
    if actual is not None:
        fig.add_vline(x=actual, line_dash='dash', line_color='gold', line_width=1.5)
        fig.add_annotation(
            x=actual, y=0.95, yref='paper',
            text="Current",
            showarrow=False,
            xanchor='left', yanchor='top',
            font=dict(color='gold', size=12))
    fig.update_layout(
        paper_bgcolor='#0E1117',
        plot_bgcolor='#0E1117',
        font_color='white',
        height=350,
        xaxis=dict(
            title=title,
            tickfont=dict(color='white'),
            gridcolor='#444444'),
        yaxis=dict(
            title='Probability (%)',
            ticksuffix='%',
            tickfont=dict(color='white'),
            gridcolor='#444444'),
        bargap=0,
        showlegend=False,
        margin=dict(l=50, r=20, t=20, b=50))
    st.plotly_chart(fig, use_container_width=True)

def score_opening_rounds(picks, simulations, projections, day="Thu"):
    round_points = 1
    r64_cols = [f'R64_{i}' for i in range(1, 33)]
    seed_map = projections.set_index("Team")["Seed"]
    day_map = projections.set_index("Team")["R64Day"]
    sim_arr = simulations[r64_cols].values
    picks_arr = picks[r64_cols].values
    seed_values = np.vectorize(lambda x: seed_map.get(x, 0))(picks_arr)
    pick_points = seed_values + round_points
    match = sim_arr[:, None, :] == picks_arr[None, :, :]
    scores = (match * pick_points[None, :, :])
    day_values = np.vectorize(lambda x: day_map.get(x, None))(picks_arr)
    day_mask = (day_values == day)
    scores = (scores * day_mask[None, :, :]).sum(axis=2)
    df = pd.DataFrame(scores, columns=picks["Bracket"].values)
    df.insert(0, "Sim", np.arange(1, len(df) + 1))
    return df

def count_opening_round_simulations(picks, simulations, projections, day="Thu"):
    r64_cols = [f'R64_{i}' for i in range(1, 33)]
    day_map = projections.set_index("Team")["R64Day"]
    picks_arr = picks[r64_cols].values
    sim_arr = simulations[r64_cols].values
    day_values = np.vectorize(lambda x: day_map.get(x, None))(picks_arr)
    day_mask = (day_values == day)
    match = sim_arr[:, None, :] == picks_arr[None, :, :]
    counts = (match * day_mask[None, :, :]).sum(axis=2)
    df = pd.DataFrame(counts, columns=picks["Bracket"].values)
    df.insert(0, "Sim", np.arange(1, len(df) + 1))
    return df

def score_simulations_by_region(picks, simulations, projections, region):
    round_points = {'R64': 1, 'R32': 4, 'S16': 8, 'E8': 16}
    round_cols = {
        'R64': [f'R64_{i}' for i in range(1, 33)],
        'R32': [f'R32_{i}' for i in range(1, 17)],
        'S16': [f'S16_{i}' for i in range(1, 9)],
        'E8':  [f'E8_{i}' for i in range(1, 5)]}
    seed_map = projections.set_index("Team")["Seed"]
    region_map = projections.set_index("Team")["Region"]
    total_scores = None
    for rnd, cols in round_cols.items():
        sim_arr = simulations[cols].values
        picks_arr = picks[cols].values
        seed_values = np.vectorize(lambda x: seed_map.get(x, 0))(picks_arr)
        region_values = np.vectorize(lambda x: region_map.get(x, None))(picks_arr)
        pick_points = seed_values + round_points[rnd]
        match = sim_arr[:, None, :] == picks_arr[None, :, :]
        region_mask = (region_values == region)
        scores = (match * pick_points[None, :, :] * region_mask[None, :, :]).sum(axis=2)
        if total_scores is None:
            total_scores = scores.copy()
        else:
            total_scores += scores
    df = pd.DataFrame(total_scores, columns=picks["Bracket"].values)
    df.insert(0, "Sim", np.arange(1, len(df) + 1))
    return df

def count_simulations_by_region(picks, simulations, projections, region):
    round_cols = {
        'R64': [f'R64_{i}' for i in range(1, 33)],
        'R32': [f'R32_{i}' for i in range(1, 17)],
        'S16': [f'S16_{i}' for i in range(1, 9)],
        'E8':  [f'E8_{i}' for i in range(1, 5)]}
    region_map = projections.set_index("Team")["Region"]
    total_counts = None
    for rnd, cols in round_cols.items():
        sim_arr = simulations[cols].values
        picks_arr = picks[cols].values
        region_values = np.vectorize(lambda x: region_map.get(x, None))(picks_arr)
        match = sim_arr[:, None, :] == picks_arr[None, :, :]
        region_mask = (region_values == region)
        counts = (match * region_mask[None, :, :]).sum(axis=2)
        if total_counts is None:
            total_counts = counts.copy()
        else:
            total_counts += counts
    df = pd.DataFrame(total_counts, columns=picks["Bracket"].values)
    df.insert(0, "Sim", np.arange(1, len(df) + 1))
    return df

def calculate_expected_value(Scores, Counts, payout, Type):
    payout = np.array(payout)
    score_vals = Scores.iloc[:, 1:].values
    count_vals = Counts.iloc[:, 1:].values
    n_sims, n_brackets = score_vals.shape
    total_returns = np.zeros(n_brackets)
    for s in range(n_sims):
        final_order = np.lexsort((-count_vals[s], -score_vals[s]))
        ranked_payout = payout[:n_brackets]
        total_returns[final_order] += ranked_payout
    expected_values = total_returns / n_sims
    return pd.DataFrame({
        "Bracket": Scores.columns[1:],
        Type: expected_values
    })

@st.cache_data
def run_analysis(sims_data, Projections, Projections2, Picks):
    Scores64, Scores32, Scores16, Scores8, Scores4, Scores2, ScoresTotal = score_simulations_by_round(Picks, sims_data, Projections)
    
    for scores in [Scores2, Scores4, Scores8, Scores16, Scores32, Scores64]:
        scores.insert(0, "Sim", range(1, len(Scores2) + 1))
    
    Counts64, Counts32, Counts16, Counts8, Counts4, Counts2, CountsTotal = count_simulations_by_round(Picks, sims_data)
    Finish = calculate_sim_ranks(ScoresTotal)
    ScoresThurs = score_opening_rounds(Picks, sims_data, Projections, "Thursday")
    ScoresFri = score_opening_rounds(Picks, sims_data, Projections, "Friday")
    CountsThurs = count_opening_round_simulations(Picks, sims_data, Projections, "Thursday")
    CountsFri = count_opening_round_simulations(Picks, sims_data, Projections, "Friday")
    ScoresWest = score_simulations_by_region(Picks, sims_data, Projections, "West")
    ScoresEast = score_simulations_by_region(Picks, sims_data, Projections, "East")
    ScoresSouth = score_simulations_by_region(Picks, sims_data, Projections, "South")
    ScoresMidwest = score_simulations_by_region(Picks, sims_data, Projections, "Midwest")
    CountsWest = count_simulations_by_region(Picks, sims_data, Projections, "West")
    CountsEast = count_simulations_by_region(Picks, sims_data, Projections, "East")
    CountsSouth = count_simulations_by_region(Picks, sims_data, Projections, "South")
    CountsMidwest = count_simulations_by_region(Picks, sims_data, Projections, "Midwest")
    TotalExpected, TotalPayout = get_total_payout(ScoresTotal, CountsTotal, Projections2, ScoresThurs, CountsThurs, sims_data, ScoresFri, CountsFri, ScoresWest, CountsWest, ScoresEast, CountsEast, ScoresSouth, CountsSouth, ScoresMidwest, CountsMidwest, Scores32, Counts32, Picks, Projections)
    ExpectedDF = get_expected_dataframe(
        ScoresTotal, CountsTotal,
        ScoresThurs, CountsThurs,
        ScoresFri, CountsFri,
        ScoresWest, CountsWest,
        ScoresEast, CountsEast,
        ScoresMidwest, CountsMidwest,
        ScoresSouth, CountsSouth,
        Scores64, Counts64,
        Scores32, Counts32,
        Scores16, Counts16,
        Scores8, Counts8,
        Scores4, Counts4,
        Scores2, Counts2
    )

    return (
        Scores64, Scores32, Scores16, Scores8, Scores4, Scores2, ScoresTotal,
        Counts64, Counts32, Counts16, Counts8, Counts4, Counts2, CountsTotal,
        Finish,
        ScoresThurs, ScoresFri, CountsThurs, CountsFri,
        ScoresWest, ScoresEast, ScoresSouth, ScoresMidwest,
        CountsWest, CountsEast, CountsSouth, CountsMidwest,
        TotalExpected, TotalPayout,
        ExpectedDF
    )

def get_expected_dataframe(
    ScoresTotal, CountsTotal,
    ScoresThurs, CountsThurs,
    ScoresFri, CountsFri,
    ScoresWest, CountsWest,
    ScoresEast, CountsEast,
    ScoresMidwest, CountsMidwest,
    ScoresSouth, CountsSouth,
    Scores64, Counts64,
    Scores32, Counts32,
    Scores16, Counts16,
    Scores8, Counts8,
    Scores4, Counts4,
    Scores2, Counts2
):
    def mean_series(df):
        return df.drop(columns=["Sim"], errors="ignore").mean(axis=0)

    scores_map = {
        "Total":        mean_series(ScoresTotal),
        "Thurs":        mean_series(ScoresThurs),
        "Fri":          mean_series(ScoresFri),
        "West":         mean_series(ScoresWest),
        "East":         mean_series(ScoresEast),
        "Midwest":      mean_series(ScoresMidwest),
        "South":        mean_series(ScoresSouth),
        "Round of 64":  mean_series(Scores64),
        "Round of 32":  mean_series(Scores32),
        "Sweet 16":     mean_series(Scores16),
        "Elite 8":      mean_series(Scores8),
        "Final Four":   mean_series(Scores4),
        "Championship": mean_series(Scores2),
    }

    counts_map = {
        "Total":        mean_series(CountsTotal),
        "Thurs":        mean_series(CountsThurs),
        "Fri":          mean_series(CountsFri),
        "West":         mean_series(CountsWest),
        "East":         mean_series(CountsEast),
        "Midwest":      mean_series(CountsMidwest),
        "South":        mean_series(CountsSouth),
        "Round of 64":  mean_series(Counts64),
        "Round of 32":  mean_series(Counts32),
        "Sweet 16":     mean_series(Counts16),
        "Elite 8":      mean_series(Counts8),
        "Final Four":   mean_series(Counts4),
        "Championship": mean_series(Counts2),
    }

    df = pd.DataFrame({"Bracket": scores_map["Total"].index})
    for col, series in scores_map.items():
        df[f"{col}_Score"] = series.values
    for col, series in counts_map.items():
        df[f"{col}_Count"] = series.values

    return df

def get_scores_dataframe(Actual, Projections, Projections2, Picks):
    Scores64, Scores32, Scores16, Scores8, Scores4, Scores2, ScoresTotal = score_simulations_by_round(Picks, Actual, Projections)
    Counts64, Counts32, Counts16, Counts8, Counts4, Counts2, CountsTotal = count_simulations_by_round(Picks, Actual)
    ScoresThurs = score_opening_rounds(Picks, Actual, Projections, "Thursday")
    ScoresFri   = score_opening_rounds(Picks, Actual, Projections, "Friday")
    CountsThurs = count_opening_round_simulations(Picks, Actual, Projections, "Thursday")
    CountsFri   = count_opening_round_simulations(Picks, Actual, Projections, "Friday")
    ScoresWest  = score_simulations_by_region(Picks, Actual, Projections, "West")
    ScoresEast  = score_simulations_by_region(Picks, Actual, Projections, "East")
    ScoresSouth = score_simulations_by_region(Picks, Actual, Projections, "South")
    ScoresMidwest = score_simulations_by_region(Picks, Actual, Projections, "Midwest")
    CountsWest  = count_simulations_by_region(Picks, Actual, Projections, "West")
    CountsEast  = count_simulations_by_region(Picks, Actual, Projections, "East")
    CountsSouth = count_simulations_by_region(Picks, Actual, Projections, "South")
    CountsMidwest = count_simulations_by_region(Picks, Actual, Projections, "Midwest")
    def extract(df):
        return df.drop(columns=["Sim"], errors="ignore").iloc[0]
    scores_map = {
        "Total":          extract(ScoresTotal),
        "Thurs":          extract(ScoresThurs),
        "Fri":            extract(ScoresFri),
        "West":           extract(ScoresWest),
        "East":           extract(ScoresEast),
        "Midwest":        extract(ScoresMidwest),
        "South":          extract(ScoresSouth),
        "Round of 64":    extract(Scores64),
        "Round of 32":    extract(Scores32),
        "Sweet 16":       extract(Scores16),
        "Elite 8":        extract(Scores8),
        "Final Four":     extract(Scores4),
        "Championship":   extract(Scores2),
    }
    counts_map = {
        "Total":          extract(CountsTotal),
        "Thurs":          extract(CountsThurs),
        "Fri":            extract(CountsFri),
        "West":           extract(CountsWest),
        "East":           extract(CountsEast),
        "Midwest":        extract(CountsMidwest),
        "South":          extract(CountsSouth),
        "Round of 64":    extract(Counts64),
        "Round of 32":    extract(Counts32),
        "Sweet 16":       extract(Counts16),
        "Elite 8":        extract(Counts8),
        "Final Four":     extract(Counts4),
        "Championship":   extract(Counts2),
    }
    brackets = scores_map["Total"].index
    df = pd.DataFrame({"Bracket": brackets})
    for col, series in scores_map.items():
        df[f"{col}_Score"] = series.values
    for col, series in counts_map.items():
        df[f"{col}_Count"] = series.values
    return df

def get_projections2():
    csv_url = "https://docs.google.com/spreadsheets/d/12f4bu9JRwZ9TDXVw6T2GI0fPgjeKCk1GxrdDFdjHds8/export?format=csv&gid=783864246"
    df = pd.read_csv(csv_url)
    df = df[df["Scheduled"]]
    df = df.drop(columns=["Scheduled"])
    return df

def build_payout_matrix(ScoresFinal, ScoresCounts, payout):
    ScoresFinal = ScoresFinal.drop(columns=["Sim"], errors="ignore")
    ScoresCounts = ScoresCounts.drop(columns=["Sim"], errors="ignore")
    payout = np.array(payout)
    scores = ScoresFinal.values
    counts = ScoresCounts.values
    n_sims, n_brackets = scores.shape
    payout_matrix = np.zeros((n_sims, n_brackets))
    for s in range(n_sims):
        order = np.lexsort((-counts[s], -scores[s]))
        payout_matrix[s, order] = payout[:n_brackets]
    return payout_matrix

def build_ev_table(Projections2, ScoresTotal, CountsTotal, simulations, payout, Type):
    ScoresTotal = ScoresTotal.drop(columns=["Sim"], errors="ignore")
    CountsTotal = CountsTotal.drop(columns=["Sim"], errors="ignore")
    PayoutMatrix = build_payout_matrix(ScoresTotal, CountsTotal, payout)
    results = []
    brackets = range(PayoutMatrix.shape[1])
    for _, game in Projections2.iterrows():
        game_id = game["GameID"]
        team_a = game["TeamA"]
        team_b = game["TeamB"]
        sim_col = simulations[game_id].values
        mask_a = sim_col == team_a
        mask_b = sim_col == team_b
        if mask_a.sum() == 0 or mask_b.sum() == 0:
            continue
        ev_a = PayoutMatrix[mask_a].mean(axis=0)
        ev_b = PayoutMatrix[mask_b].mean(axis=0)
        ev_diff = ev_a - ev_b
        for j in brackets:
            results.append([
                game_id,
                PayoutMatrix.shape[1] and j,
                ev_a[j],
                ev_b[j],
                ev_diff[j],
                Type])
    return pd.DataFrame(
        results,
        columns=["GameID", "Bracket", "EV_A", "EV_B", "EV_Diff", "Type"])

def get_total_payout(ScoresTotal, CountsTotal, Projections2, ScoresThurs, CountsThurs, Sims, ScoresFri, CountsFri, ScoresWest, CountsWest, ScoresEast, CountsEast, ScoresSouth, CountsSouth, ScoresMidwest, CountsMidwest, Scores32, Counts32, Picks, Projections):
    payout = [1400, 250, 175, 100, 50] + [20] * 33 + [0] * 161 + [5]
    ExpTotal = calculate_expected_value(ScoresTotal, CountsTotal, payout, "Total")
    TotalPayoutOutput = build_ev_table(Projections2, ScoresTotal, CountsTotal, Sims, payout, "Total")
    payout = [15] * 1 + [0] * 199
    ExpThurs = calculate_expected_value(ScoresThurs, CountsThurs, payout, "Thurs")
    ThursPayoutOutput = build_ev_table(Projections2, ScoresThurs, CountsThurs, Sims, payout, "Thurs")
    payout = [15] * 5 + [0] * 195
    ExpFri = calculate_expected_value(ScoresFri, CountsFri, payout, "Fri")
    FriPayoutOutput = build_ev_table(Projections2, ScoresFri, CountsFri, Sims, payout, "Fri")
    payout = [50] * 1 + [0] * 199
    ExpWest = calculate_expected_value(ScoresWest, CountsWest, payout, "West")
    WestPayoutOutput = build_ev_table(Projections2, ScoresWest, CountsWest, Sims, payout, "West")
    payout = [50] * 1 + [0] * 199
    ExpEast = calculate_expected_value(ScoresEast, CountsEast, payout, "East")
    EastPayoutOutput = build_ev_table(Projections2, ScoresEast, CountsEast, Sims, payout, "East")
    payout = [50] * 1 + [0] * 199
    ExpSouth = calculate_expected_value(ScoresSouth, CountsSouth, payout, "South")
    SouthPayoutOutput = build_ev_table(Projections2, ScoresSouth, CountsSouth, Sims, payout, "South")
    payout = [50] * 1 + [0] * 199
    ExpMidwest = calculate_expected_value(ScoresMidwest, CountsMidwest, payout, "Midwest")
    MidwestPayoutOutput = build_ev_table(Projections2, ScoresMidwest, CountsMidwest, Sims, payout, "Midwest")
    payout = [25] * 1 + [0] * 199
    ExpS16 = calculate_expected_value(Scores32, Counts32, payout, "S16")
    S16PayoutOutput = build_ev_table(Projections2, Scores32, Counts32, Sims, payout, "S16")
    exp_tables = [
        ExpTotal, ExpThurs, ExpFri,
        ExpWest, ExpEast, ExpSouth,
        ExpMidwest, ExpS16]
    ExpCombined = reduce(
        lambda left, right: pd.merge(left, right, on="Bracket", how="left"),
        exp_tables)
    payout_tables = [
        TotalPayoutOutput, ThursPayoutOutput, FriPayoutOutput,
        WestPayoutOutput, EastPayoutOutput, SouthPayoutOutput,
        MidwestPayoutOutput, S16PayoutOutput]
    PayoutCombined = pd.concat(payout_tables, axis=0, ignore_index=True)
    PayoutCombined = (
    PayoutCombined
    .groupby(["GameID", "Bracket"], as_index=False)
    .agg({
        "EV_A": "sum",
        "EV_B": "sum"}))
    PayoutCombined["EV_Diff"] = PayoutCombined["EV_B"] - PayoutCombined["EV_A"]
    #PayoutCombined["Bracket"] = np.tile(Picks["Bracket"].values, len(PayoutCombined) // len(Picks))
    score_columns = ScoresTotal.columns[1:]  # skip the Sim column
    bracket_names = pd.DataFrame({
        "Bracket": range(len(score_columns)),
        "BracketName": score_columns
    })
    PayoutCombined = PayoutCombined.merge(bracket_names, on="Bracket", how="left")
    PayoutCombined = PayoutCombined.drop(columns="Bracket").rename(columns={"BracketName": "Bracket"})
    PayoutCombined = PayoutCombined.merge(Projections2[["GameID", "TeamA", "TeamB"]], on="GameID", how="left")
    PayoutCombined = PayoutCombined.merge(Projections[["Team", "Seed", "ActualName", "Logo", "Record", "Color"]], left_on="TeamA", right_on="Team", how="left", suffixes=("", "_a")).drop(columns="Team")
    PayoutCombined = PayoutCombined.merge(Projections[["Team", "Seed", "ActualName", "Logo", "Record", "Color"]], left_on="TeamB", right_on="Team", how="left", suffixes=("", "_b")).drop(columns="Team")
    PayoutCombined["Record"] = ("No. " + PayoutCombined["Seed"].astype(str) + " " + PayoutCombined["Record"])
    PayoutCombined["Record_b"] = ("No. " + PayoutCombined["Seed_b"].astype(str) + " " + PayoutCombined["Record_b"])
    PayoutCombined = PayoutCombined[["Bracket",  "ActualName", "ActualName_b", "Logo", "Logo_b", "Record", "Record_b", "Color", "Color_b", "EV_A", "EV_B", "EV_Diff"]]
    return ExpCombined, PayoutCombined


def render_ev_matchup(df, RowNumber):

    row = df.iloc[RowNumber]

    team_a   = row["ActualName"]
    team_b   = row["ActualName_b"]
    logo_a   = row["Logo"]
    logo_b   = row["Logo_b"]
    record_a = row["Record"]
    record_b = row["Record_b"]
    color_a  = row["Color"]
    color_b  = row["Color_b"]
    ev_a     = row["EV_A"]
    ev_b     = row["EV_B"]
    ev_diff  = row["EV_Diff"]

    # ── normalize indicator position ─────────────────────────────────────────
    # Map ev_diff onto [-1, 1] using the sum of absolute EVs as scale.
    # Clamp to avoid extreme edge cases.
    scale = abs(ev_a) + abs(ev_b)
    if scale == 0:
        norm = 0.0
    else:
        norm = max(-1.0, min(1.0, ev_diff / scale))

    # Convert norm [-1,1] → bar percentage [0,100], center = 50
    indicator_pct = 50 + norm * 45  # max 5% padding from edges

    # ── logo src helper ───────────────────────────────────────────────────────
    def logo_src(path):
        if path.startswith("http"):
            return path
        import base64 
        import pathlib
        data = pathlib.Path(path).read_bytes()
        ext = pathlib.Path(path).suffix.lstrip(".")
        mime = "png" if ext == "png" else "jpeg" if ext in ("jpg", "jpeg") else "png"
        return f"data:image/{mime};base64,{base64.b64encode(data).decode()}"

    src_a = logo_src(logo_a)
    src_b = logo_src(logo_b) ##A

    # ── dollar formatting ─────────────────────────────────────────────────────
    def fmt(v):
        return f"${abs(v):.2f}"

    label_a = fmt(ev_a)
    label_b = fmt(ev_b)

    # Which side has the edge?
    edge_side = "A" if ev_diff > 0 else "B" if ev_diff < 0 else None
    edge_color = color_b if edge_side == "A" else color_a if edge_side == "B" else "#ffffff"

    html = f"""
    <link href="https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;600;700;900&family=Barlow:wght@400;500;600&display=swap" rel="stylesheet">

    <style>
      .ev-root {{
        font-family: 'Barlow', sans-serif;
        background: #0a0a0f;
        border-radius: 16px;
        overflow: visible;
        position: relative;
        box-shadow:
          0 0 0 1px rgba(255,255,255,0.06),
          0 32px 80px rgba(0,0,0,0.7),
          0 8px 24px rgba(0,0,0,0.5);
        margin-bottom: 8px;
        padding: 20px 24px 28px;
      }}

      /* grain overlay */
      .ev-root::before {{
        content: '';
        position: absolute;
        inset: 0;
        background-image: url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)' opacity='0.04'/%3E%3C/svg%3E");
        opacity: 0.35;
        pointer-events: none;
        z-index: 0;
        border-radius: 16px;
      }}

      /* split background glow */
      .ev-glow-left {{
        position: absolute; left: -5%; top: -20%;
        width: 50%; height: 140%;
        background: radial-gradient(ellipse at 20% 50%, {color_a}55 0%, transparent 60%);
        pointer-events: none; z-index: 0; filter: blur(4px);
      }}
      .ev-glow-right {{
        position: absolute; right: -5%; top: -20%;
        width: 50%; height: 140%;
        background: radial-gradient(ellipse at 80% 50%, {color_b}55 0%, transparent 60%);
        pointer-events: none; z-index: 0; filter: blur(4px);
      }}

      /* teams row */
      .ev-teams {{
        position: relative; z-index: 5;
        display: flex;
        align-items: center;
        justify-content: space-between;
        margin-bottom: 22px;
      }}

      .ev-team {{
        display: flex;
        align-items: center;
        gap: 14px;
        flex: 1;
      }}
      .ev-team.right {{
        flex-direction: row-reverse;
        text-align: right;
      }}

      .ev-logo {{
        width: 72px; height: 72px;
        object-fit: contain;
        flex-shrink: 0;
        filter: drop-shadow(0 4px 16px rgba(0,0,0,0.6));
      }}

      .ev-team-info {{ line-height: 1.2; }}

      .ev-team-name {{
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 26px; font-weight: 900;
        text-transform: uppercase;
        color: #ffffff;
        letter-spacing: 0.5px;
        display: block; line-height: 1;
      }}
      .ev-team-record {{
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 12px; font-weight: 600;
        letter-spacing: 2px;
        color: rgba(255,255,255,0.35);
        display: block; margin-top: 4px;
        text-transform: uppercase;
      }}

      /* VS badge */
      .ev-vs {{
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 13px; font-weight: 700;
        letter-spacing: 3px;
        color: rgba(255,255,255,0.2);
        padding: 0 20px;
        flex-shrink: 0;
      }}

      /* ── mobile: stack teams vertically ── */
      @media (max-width: 520px) {{
        .ev-teams {{
          flex-direction: column;
          gap: 0;
          margin-bottom: 16px;
        }}
        .ev-team {{
          flex-direction: row !important;
          text-align: left !important;
          width: 100%;
          padding: 10px 0;
        }}
        .ev-team.right {{
          flex-direction: row-reverse !important;
          text-align: right !important;
          border-top: 1px solid rgba(255,255,255,0.07);
        }}
        .ev-logo {{
          width: 52px; height: 52px;
        }}
        .ev-team-name {{
          font-size: 20px;
        }}
        .ev-vs {{
          display: none;
        }}
      }}

      /* ── EV bar section ── */
      .ev-bar-section {{
        position: relative; z-index: 5;
        margin-top: 8px;
      }}

      /* label row above bar */
      .ev-labels {{
        display: flex;
        justify-content: space-between;
        align-items: flex-end;
        margin-bottom: 10px;
      }}
      .ev-label {{
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 11px; font-weight: 700;
        letter-spacing: 2px;
        text-transform: uppercase;
        color: rgba(255,255,255,0.4);
      }}
      .ev-amount {{
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 22px; font-weight: 900;
        letter-spacing: -0.5px;
        display: block; line-height: 1;
      }}
      .ev-amount.a {{ color: #ffffff; text-shadow: 0 0 20px rgba(255,255,255,0.2); }}
      .ev-amount.b {{ color: #ffffff; text-shadow: 0 0 20px rgba(255,255,255,0.2); }}

      .ev-label-block {{ display: flex; flex-direction: column; gap: 3px; }}
      .ev-label-block.right {{ text-align: right; }}

      /* the track */
      .ev-track {{
        position: relative;
        height: 8px;
        border-radius: 99px;
        background: rgba(255,255,255,0.08);
        overflow: visible;
      }}

      /* left fill (Team A color) */
      .ev-fill-a {{
        position: absolute;
        left: 0; top: 0; bottom: 0;
        width: {indicator_pct}%;
        border-radius: 99px 0 0 99px;
        background: linear-gradient(90deg, {color_a}99, {color_a}dd);
        transition: width 0.6s cubic-bezier(0.34, 1.56, 0.64, 1);
      }}

      /* right fill (Team B color) */
      .ev-fill-b {{
        position: absolute;
        right: 0; top: 0; bottom: 0;
        width: {100 - indicator_pct}%;
        border-radius: 0 99px 99px 0;
        background: linear-gradient(270deg, {color_b}99, {color_b}dd);
        transition: width 0.6s cubic-bezier(0.34, 1.56, 0.64, 1);
      }}

      /* glowing indicator dot */
      .ev-indicator {{
        position: absolute;
        top: 50%;
        left: {indicator_pct}%;
        transform: translate(-50%, -50%);
        width: 20px; height: 20px;
        border-radius: 50%;
        background: {edge_color};
        box-shadow:
          0 0 0 3px rgba(10,10,15,0.9),
          0 0 0 5px {edge_color}66,
          0 0 20px 6px {edge_color}88,
          0 0 40px 12px {edge_color}44;
        z-index: 10;
        animation: indicatorPulse 2s ease-in-out infinite;
      }}

      @keyframes indicatorPulse {{
        0%, 100% {{ box-shadow:
          0 0 0 3px rgba(10,10,15,0.9),
          0 0 0 5px {edge_color}66,
          0 0 20px 6px {edge_color}88,
          0 0 40px 12px {edge_color}44; }}
        50% {{ box-shadow:
          0 0 0 3px rgba(10,10,15,0.9),
          0 0 0 6px {edge_color}99,
          0 0 30px 10px {edge_color}aa,
          0 0 60px 20px {edge_color}55; }}
      }}

      .ev-callout {{
        position: absolute;
        left: {indicator_pct}%;
        bottom: 100%;
        transform: translateX(-50%);
        display: flex;
        flex-direction: column;
        align-items: center;
        margin-bottom: 6px;
        pointer-events: none;
      }}
      .ev-callout-bubble {{
        background: rgba(255,255,255,0.12);
        border: 1px solid rgba(255,255,255,0.2);
        color: #ffffff;
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 13px; font-weight: 700;
        letter-spacing: 0.5px;
        padding: 4px 12px;
        border-radius: 6px;
        white-space: nowrap;
        margin-bottom: 4px;
      }}
      .ev-callout-line {{
        width: 2px;
        height: 10px;
        background: linear-gradient(rgba(255,255,255,0.4), transparent);
        border-radius: 1px;
      }}

      /* bottom team name labels under bar */
      .ev-bar-labels {{
        display: flex;
        justify-content: space-between;
        margin-top: 10px;
      }}
      .ev-bar-team-label {{
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 10px; font-weight: 700;
        letter-spacing: 2.5px;
        text-transform: uppercase;
        color: rgba(255,255,255,0.25);
      }}

      /* color accent stripes on the card edges */
      .ev-stripe-a {{
        position: absolute; left: 0; top: 0; bottom: 0;
        width: 4px; border-radius: 16px 0 0 16px;
        background: {color_a};
        box-shadow: 2px 0 16px {color_a}88;
      }}
      .ev-stripe-b {{
        position: absolute; right: 0; top: 0; bottom: 0;
        width: 4px; border-radius: 0 16px 16px 0;
        background: {color_b};
        box-shadow: -2px 0 16px {color_b}88;
      }}
    </style>

    <div class="ev-root">
      <div class="ev-stripe-a"></div>
      <div class="ev-stripe-b"></div>
      <div class="ev-glow-left"></div>
      <div class="ev-glow-right"></div>

      <!-- Teams -->
      <div class="ev-teams">
        <div class="ev-team left">
          <img class="ev-logo" src="{src_a}" alt="{team_a}" onerror="this.style.opacity='0.3'" />
          <div class="ev-team-info">
            <span class="ev-team-name">{team_a}</span>
            <span class="ev-team-record">{record_a}</span>
          </div>
        </div>

        <div class="ev-vs">VS</div>

        <div class="ev-team right">
          <img class="ev-logo" src="{src_b}" alt="{team_b}" onerror="this.style.opacity='0.3'" />
          <div class="ev-team-info">
            <span class="ev-team-name">{team_b}</span>
            <span class="ev-team-record">{record_b}</span>
          </div>
        </div>
      </div>

      <!-- EV Bar -->
      <div class="ev-bar-section">
        <!-- Dollar labels above bar -->
        <div class="ev-labels">
          <div class="ev-label-block">
            <span class="ev-label">If {team_a} wins</span>
            <span class="ev-amount a">{label_a}</span>
          </div>
          <div class="ev-label-block right">
            <span class="ev-label">If {team_b} wins</span>
            <span class="ev-amount b">{label_b}</span>
          </div>
        </div>

        <!-- Track with fills and indicator -->
        <div style="position:relative; padding-top: 40px;">
          <!-- Callout above indicator -->
          <div class="ev-callout">
            <div class="ev-callout-bubble">
                +${abs(ev_diff):.2f}
            </div>
            <div class="ev-callout-line"></div>
          </div>

          <div class="ev-track">
            <div class="ev-fill-a"></div>
            <div class="ev-fill-b"></div>
            <div class="ev-indicator"></div>
          </div>
        </div>

        <!-- Team name labels under bar -->
        <div class="ev-bar-labels">
          <span class="ev-bar-team-label">◀ {team_a}</span>
          <span class="ev-bar-team-label">{team_b} ▶</span>
        </div>
      </div>
    </div>
    """

    components.html(html, height=380, scrolling=False)

def update_total_expected(TotalExpected, ScoresTotal, CountTotal, RiskScore, Picks, Finish, ActualResultsExp):
    df = TotalExpected.copy()
    sum_cols = ["Total", "Thurs", "Fri", "West", "East", "South", "Midwest", "S16"]
    df["Total"] = df[sum_cols].sum(axis=1)
    scale_cols2 = ["Thurs", "Fri"]
    for col in scale_cols2:
        df[col] = df[col] * 100 / 15
    scale_cols = ["West", "East", "South", "Midwest", "S16"]
    for col in scale_cols:
        df[col] = df[col] * 100 / 50
    pred_correct = ScoresTotal.mean(axis=0).reset_index()
    pred_correct.columns = ["Bracket", "Predicted_Correct_Picks"]
    pred_points = CountTotal.mean(axis=0).reset_index()
    pred_points.columns = ["Bracket", "Predicted_Points"]
    df = df.merge(pred_correct, on="Bracket", how="left")
    df = df.merge(pred_points, on="Bracket", how="left")
    risk_cols = ["risk_score", "downside", "champ_concentration", "avg_upset_seed"]
    df = df.merge( RiskScore[["Bracket"] + risk_cols], on="Bracket", how="left")
    picks = Picks.copy()
    picks["Runner_Up"] = picks.apply(lambda row: row["F4_1"] if row["F4_1"] != row["Champ"] else row["F4_2"], axis=1)
    df = df.merge(picks[["Bracket", "Champ", "Runner_Up"]], on="Bracket", how="left")
    finish = Finish.drop(columns=["Sim"], errors="ignore")
    n_sims = finish.shape[0]
    win_pct = ((finish == 1).sum(axis=0).div(n_sims).mul(100).reset_index())
    win_pct.columns = ["Bracket", "Win %"]
    money_pct = ((finish <= 40).sum(axis=0).div(n_sims).mul(100).reset_index())
    money_pct.columns = ["Bracket", "In The Money %"]
    df = df.merge(win_pct, on="Bracket", how="left")
    df = df.merge(money_pct, on="Bracket", how="left")
    df = df.merge(ActualResultsExp[["Bracket", "Total_Score"]], on="Bracket", how="left")
    rename_map = {
        "Total": "Total EV",
        "Total_Score": "Score",
        "Thurs": "Th Win%",
        "Fri": "Fr Win%",
        "West": "W Win%",
        "East": "E Win%",
        "South": "S Win%",
        "Midwest": "MW Win%",
        "S16": "S16 Win%",
        "Predicted_Correct_Picks": "Pred. Pts",
        "Predicted_Points": "Pred. Games",
        "risk_score": "Risk Score",
        "downside": "Downside",
        "champ_concentration": "Champ Risk",
        "avg_upset_seed": "Avg. Upset",
        "Win %": "Win%",
        "In The Money %": "ITM%",
        "Champ": "Champion",
        "Runner_Up": "Runner Up"
    } #A
    df = df.rename(columns=rename_map)
    ordered_cols = [
        "Bracket",
        "Score",
        "Pred. Pts",
        "Pred. Games",
        "Total EV",
        "Win%",
        "ITM%",
        "Th Win%",
        "Fr Win%",
        "W Win%",
        "E Win%",
        "S Win%",
        "MW Win%",
        "S16 Win%",
        "Risk Score",
        "Downside",
        "Champ Risk",
        "Avg. Upset",
        "Champion",
        "Runner Up"
    ]
    ordered_cols = [c for c in ordered_cols if c in df.columns]
    df = df[ordered_cols]
    return df

"""
March Madness Bracket Renderer for Streamlit
=============================================

Usage
-----
    from march_madness_bracket import render_bracket

    # projections: 64-row DataFrame
    #   Required columns: Region, Seed, ActualName, Color, Logo, Record
    #   - Region   : one of 4 region names (consistent order matters – see below)
    #   - Seed     : 1-16
    #   - ActualName: display name of the team
    #   - Color    : hex color string, e.g. '#1a4789'
    #   - Logo     : URL string (or empty)
    #   - Record   : e.g. '28-5'

    # picks: DataFrame with one row per bracket
    #   Required columns:
    #     Bracket                 – bracket owner/name
    #     R64_1 … R64_32         – winners of Round of 64 (team names)
    #     R32_1 … R32_16         – winners of Round of 32
    #     S16_1 … S16_8          – winners of Sweet 16
    #     E8_1  … E8_4           – winners of Elite 8
    #     F4_1, F4_2             – winners of Final Four games
    #     Champ                  – champion

    # Game-numbering convention (matches your column ordering):
    #   The 4 regions appear in the order projections['Region'].unique() returns them.
    #   Region 0 → R64_1..8,  R32_1..4,  S16_1..2, E8_1
    #   Region 1 → R64_9..16, R32_5..8,  S16_3..4, E8_2
    #   Region 2 → R64_17..24, R32_9..12, S16_5..6, E8_3
    #   Region 3 → R64_25..32, R32_13..16, S16_7..8, E8_4
    #   F4_1 = left-side F4 (E8_1 winner vs E8_2 winner)
    #   F4_2 = right-side F4 (E8_3 winner vs E8_4 winner)

    render_bracket(projections, picks)          # shows bracket selector
    render_bracket(projections, picks, "Alice") # pre-selects Alice's bracket
"""

# ── Seeding order per region (top → bottom) ──────────────────────────────────
SEED_ORDER = [1, 16, 9, 8, 5, 12, 13, 4, 3, 14, 11, 6, 7, 10, 15, 2]

# ── Layout constants (pixels) ─────────────────────────────────────────────────
SLOT_H   = 26      # height of one team slot
W        = dict(R64=132, R32=116, S16=100, E8=88, F4=124)
CONN     = 10      # connector extension on each slot (bracket line space)
LBL_H   = 18      # height of round-label row
REG_GAP = 16      # vertical gap between the two regions on each side

# Bracket-line spacers in SLOT_H units: (before-first-matchup, between-matchups)
SPACERS = dict(R64=(0, 0), R32=(1, 2), S16=(3, 6), E8=(7, 0))

BL  = '#009CDE'    # bracket line colour
BG  = '#0e1117'    # page background


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _safe(val):
    """Return stripped string or None."""
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except Exception:
        pass
    s = str(val).strip()
    return s if s else None


def _build_lookups(proj: pd.DataFrame):
    by_name = {}
    by_rs   = {}   # (region_str, seed_int) → team dict
    for _, row in proj.iterrows():
        name = _safe(row.get("ActualName"))
        if not name:
            continue
        t = {
            "name":   name,
            "seed":   int(row["Seed"]),
            "color":  _safe(row.get("Color"))  or "#334155",
            "logo":   _safe(row.get("Logo"))   or "",
            "record": _safe(row.get("Record")) or "",
        }
        by_name[name] = t
        by_rs[(str(row["Region"]), int(row["Seed"]))] = t
    return by_name, by_rs


# ─────────────────────────────────────────────────────────────────────────────
# Single team-slot renderer
# ─────────────────────────────────────────────────────────────────────────────

def _slot(name, by_name, is_winner, pos, side, width):
    """
    Render one team slot div.

    Parameters
    ----------
    name     : team name (str) or None → empty grey slot
    by_name  : dict from _build_lookups
    is_winner: whether this team won the game shown in this column
    pos      : 'top' | 'bot'  – determines which bracket-line border
    side     : 'left' | 'right' – determines which side the vertical line is on
    width    : inner team-display width (px); total div = width + CONN
    """
    total = width + CONN
    v_border = f'border-{"right" if side == "left" else "left"}:1px solid {BL};'
    h_border = f'border-{"top"   if pos  == "top"  else "bottom"}:1px solid {BL};'

    if not name or name not in by_name:
        return (
            f'<div style="height:{SLOT_H}px;width:{total}px;'
            f'background:{BG};{v_border}{h_border}box-sizing:border-box;'
            f'flex-shrink:0;"></div>'
        )

    t     = by_name[name]
    color = t["color"]
    seed  = t["seed"]
    logo  = t["logo"]

    ibg    = "#14203d" if is_winner else "#0c1123"
    nc     = "#edf2f7" if is_winner else "#5a6e8c"
    sc     = "#94a3b8" if is_winner else "#3a4d6a"
    glow   = f"box-shadow:inset 0 0 0 1px {color}66;" if is_winner else ""
    ibord  = f"1px solid {color}44" if is_winner else "1px solid #18253a"
    bar_op = "0.9" if is_winner else "0.38"

    logo_h = ""
    if logo:
        logo_h = (
            f'<img src="{logo}" '
            f'style="width:13px;height:13px;object-fit:contain;flex-shrink:0;" '
            f'onerror="this.style.display=\'none\'">'
        )

    mc = max(6, (width - 50) // 7)
    dn = (name[:mc] + "…") if len(name) > mc else name

    bar = (
        f'<div style="position:absolute;left:0;top:0;bottom:0;'
        f'width:3px;background:{color};opacity:{bar_op};"></div>'
    )

    inner = (
        f'<div style="position:relative;width:{width}px;height:{SLOT_H}px;'
        f'background:{ibg};border:{ibord};{glow}'
        f'display:flex;align-items:center;padding:0 4px 0 7px;gap:3px;'
        f'overflow:hidden;box-sizing:border-box;flex-shrink:0;">'
        f'{bar}'
        f'<span style="font-size:9px;font-weight:700;color:{sc};'
        f'min-width:12px;text-align:right;flex-shrink:0;">{seed}</span>'
        f'{logo_h}'
        f'<span style="font-size:11px;font-weight:600;color:{nc};'
        f'overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">{dn}</span>'
        f'</div>'
    )

    if side == "right":
        inner = f'<div style="flex:1;min-width:{CONN}px;"></div>' + inner

    return (
        f'<div style="height:{SLOT_H}px;width:{total}px;'
        f'display:flex;align-items:center;'
        f'{v_border}{h_border}box-sizing:border-box;flex-shrink:0;">'
        f'{inner}</div>'
    )


# ─────────────────────────────────────────────────────────────────────────────
# Round-column renderer
# ─────────────────────────────────────────────────────────────────────────────


def _rcol(label, matchups, round_key, side, width, by_name):
    total    = width + CONN
    pre_u, inter_u = SPACERS[round_key]
    region_h = 16 * SLOT_H

    rows_html = ""
    for i, (n1, n2, winner) in enumerate(matchups):
        sp_u = pre_u if i == 0 else inter_u
        if sp_u > 0:
            rows_html += (
                f'<div style="height:{sp_u * SLOT_H}px;width:{total}px;'
                f'flex-shrink:0;background:{BG};"></div>'
            )
        w1 = bool(n1 and winner and n1 == winner)
        w2 = bool(n2 and winner and n2 == winner)
        rows_html += _slot(n1, by_name, w1, "top", side, width)
        rows_html += _slot(n2, by_name, w2, "bot", side, width)

    lbl = (
        f'<div style="height:{LBL_H}px;width:{total}px;'
        f'display:flex;align-items:flex-end;justify-content:center;'
        f'font-size:13px;letter-spacing:1.5px;text-transform:uppercase;'
        f'color:#2e4060;padding-bottom:2px;">{label}</div>'
    )

    body = (
        f'<div style="height:{region_h}px;width:{total}px;'
        f'display:flex;flex-direction:column;overflow:hidden;">'
        f'{rows_html}</div>'
    )

    return f'<div style="display:flex;flex-direction:column;">{lbl}{body}</div>'


# ─────────────────────────────────────────────────────────────────────────────
# Region data builder
# ─────────────────────────────────────────────────────────────────────────────

def _region_data(proj, regions, region_idx, by_name, by_rs, picks_row):
    """Return (region_name, r64, r32, s16, e8) where each is list of (n1,n2,winner)."""
    region = regions[region_idx]
    r64o   = region_idx * 8
    r32o   = region_idx * 4
    s16o   = region_idx * 2

    def gp(col):
        return _safe(picks_row.get(col, None) if hasattr(picks_row, "get")
                     else (picks_row[col] if col in picks_row.index else None))

    # R64
    r64 = []
    for i in range(8):
        s1 = SEED_ORDER[i * 2]
        s2 = SEED_ORDER[i * 2 + 1]
        t1 = by_rs.get((region, s1), {}).get("name")
        t2 = by_rs.get((region, s2), {}).get("name")
        col = r64o + i + 1
        if i == 4: 
            col = r64o + 6
        elif i == 5: 
            col = r64o + 5
        r64.append((t1, t2, gp(f"R64_{col}")))

    # R32
    r32 = []
    for i in range(4):
        # Bottom half of region: games 5&6 are swapped vs visual order
        if i == 2:
            top_col = f"R64_{r64o + i * 2 + 2}"
            bot_col = f"R64_{r64o + i * 2 + 1}"
        else:
            top_col = f"R64_{r64o + i * 2 + 1}"
            bot_col = f"R64_{r64o + i * 2 + 2}"
        r32.append((gp(top_col), gp(bot_col), gp(f"R32_{r32o + i + 1}")))

    # S16
    s16 = []
    for i in range(2):
        s16.append((
            gp(f"R32_{r32o + i * 2 + 1}"),
            gp(f"R32_{r32o + i * 2 + 2}"),
            gp(f"S16_{s16o + i + 1}"),
        ))

    # E8
    e8 = [(
        gp(f"S16_{s16o + 1}"),
        gp(f"S16_{s16o + 2}"),
        gp(f"E8_{region_idx + 1}"),
    )]

    return region, r64, r32, s16, e8


# ─────────────────────────────────────────────────────────────────────────────
# Region HTML
# ─────────────────────────────────────────────────────────────────────────────

def _region_html(proj, regions, region_idx, by_name, by_rs, picks_row, side):
    region_name, r64, r32, s16, e8 = _region_data(
        proj, regions, region_idx, by_name, by_rs, picks_row
    )

    cols = {
    "R64": _rcol("Round of 64", r64, "R64", side, W["R64"], by_name),
    "R32": _rcol("Round of 32", r32, "R32", side, W["R32"], by_name),
    "S16": _rcol("Sweet Sixteen",    s16, "S16", side, W["S16"], by_name),
    "E8":  _rcol("Elite Eight", e8,  "E8",  side, W["E8"],  by_name),
    }

    order = ["R64", "R32", "S16", "E8"] if side == "left" else ["E8", "S16", "R32", "R64"]
    rounds_html = "".join(cols[k] for k in order)

    lbl = (
        f'<div style="font-size:16px;font-weight:700;letter-spacing:3px;'
        f'text-transform:uppercase;color:#e8c96a;padding:0 4px 4px 4px;">'
        f'{region_name}</div>'
    )

    return f'<div>{lbl}<div style="display:flex;">{rounds_html}</div></div>'


# ─────────────────────────────────────────────────────────────────────────────
# Center (Final Four + Champion)
# ─────────────────────────────────────────────────────────────────────────────

def _f4_slot(name, by_name, is_winner):
    if not name:
        return (
            f'<div style="height:{SLOT_H}px;width:{W["F4"]}px;'
            f'background:#0c1123;border:1px solid #1a2840;'
            f'box-sizing:border-box;margin:1px 0;"></div>'
        )
    t     = by_name.get(name, {"name": name, "seed": "?", "color": "#334", "logo": ""})
    color = t["color"]
    ibg   = "#14203d" if is_winner else "#0c1123"
    nc    = "#edf2f7" if is_winner else "#5a6e8c"
    sc    = "#94a3b8" if is_winner else "#3a4d6a"
    logo  = t.get("logo", "")
    seed  = t.get("seed", "?")
    glow  = f"box-shadow:inset 0 0 0 1px {color}66;" if is_winner else ""
    ibord = f"1px solid {color}44" if is_winner else "1px solid #18253a"
    bar_op = "0.9" if is_winner else "0.38"
    bar = (
        f'<div style="position:absolute;left:0;top:0;bottom:0;'
        f'width:3px;background:{color};opacity:{bar_op};"></div>'
    )
    logo_h = (
        f'<img src="{logo}" style="width:13px;height:13px;object-fit:contain;" '
        f'onerror="this.style.display=\'none\'">'
    ) if logo else ""
    mc = max(8, (W["F4"] - 50) // 7)
    dn = (name[:mc] + "…") if len(name) > mc else name
    return (
        f'<div style="position:relative;height:{SLOT_H}px;width:{W["F4"]}px;'
        f'background:{ibg};border:{ibord};{glow}'
        f'display:flex;align-items:center;padding:0 5px 0 8px;gap:3px;'
        f'overflow:hidden;box-sizing:border-box;flex-shrink:0;margin:1px 0;">'
        f'{bar}'
        f'<span style="font-size:9px;font-weight:700;color:{sc};min-width:12px;">{seed}</span>'
        f'{logo_h}'
        f'<span style="font-size:11px;font-weight:600;color:{nc};'
        f'overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">{dn}</span>'
        f'</div>'
    )


def _f4_block(t1, t2, winner, label, by_name):
    w1 = bool(t1 and winner and t1 == winner)
    w2 = bool(t2 and winner and t2 == winner)
    lbl = (
        f'<div style="font-size:13px;letter-spacing:2px;text-transform:uppercase;'
        f'color:#2e4060;text-align:center;margin-bottom:2px;">{label}</div>'
    )
    return f'<div style="margin:10px 0 4px;">{lbl}{_f4_slot(t1, by_name, w1)}{_f4_slot(t2, by_name, w2)}</div>'


def _center_html(picks_row, by_name):
    def gp(col):
        return _safe(picks_row.get(col, None) if hasattr(picks_row, "get")
                     else (picks_row[col] if col in picks_row.index else None))

    f4_1_t1 = gp("E8_1")
    f4_1_t2 = gp("E8_2")  
    f4_1 = gp("F4_1")
    f4_2_t1 = gp("E8_3")
    f4_2_t2 = gp("E8_4")
    f4_2 = gp("F4_2")
    champ   = gp("Champ")

    # Championship game participants = F4 winners
    champ_t1 = f4_1
    champ_t2 = f4_2

    # Champion box
    if champ:
        t     = by_name.get(champ, {"name": champ, "seed": "?", "record": "", "color": "#e8c96a", "logo": ""})
        rec   = t.get("record", "")
        seed  = t.get("seed", "?")
        logo  = t.get("logo", "")
        logo_h = (
            f'<img src="{logo}" style="width:22px;height:22px;'
            f'object-fit:contain;display:block;margin:0 auto 4px;" '
            f'onerror="this.style.display=\'none\'">'
        ) if logo else ""
        rec_html = f'<div style="font-size:13px;color:#6272a4;margin-top:2px;">#{seed} · {rec}</div>' if rec else ""
        body = (
            f'{logo_h}'
            f'<div style="font-size:15px;font-weight:700;color:#e8c96a;'
            f'letter-spacing:0.5px;">{t["name"]}</div>'
            f'{rec_html}'
        )
    else:
        body = '<div style="font-size:11px;color:#2e4060;letter-spacing:2px;">TBD</div>'

    champ_box = (
        f'<div style="margin:14px 0;text-align:center;">'
        f'<div style="font-size:13px;letter-spacing:3px;text-transform:uppercase;'
        f'color:#e8c96a;margin-bottom:6px;">Champion</div>'
        f'<div style="font-size:20px;margin-bottom:5px;">🏆</div>'
        f'<div style="display:inline-block;width:152px;padding:10px 12px;'
        f'background:linear-gradient(135deg,#141e3c,#090e1e);'
        f'border:2px solid #e8c96a;border-radius:3px;'
        f'box-shadow:0 0 24px #e8c96a22;box-sizing:border-box;">'
        f'{body}</div></div>'
    )

    champ_game = (
        f'<div style="margin:4px 0;">'
        f'<div style="font-size:13px;letter-spacing:2px;text-transform:uppercase;'
        f'color:#2e4060;text-align:center;margin-bottom:2px;">Championship</div>'
        f'{_f4_slot(champ_t1, by_name, bool(champ_t1 and champ_t1 == champ))}'
        f'{_f4_slot(champ_t2, by_name, bool(champ_t2 and champ_t2 == champ))}'
        f'</div>'
    )

    f4_w = W["F4"] + 24
    return (
        f'<div style="display:flex;flex-direction:column;align-items:center;'
        f'padding:0 10px;min-width:{f4_w}px;">'
        f'{champ_box}'
        f'{champ_game}'
        f'{_f4_block(f4_1_t1, f4_1_t2, f4_1, "Final Four", by_name)}'
        f'{_f4_block(f4_2_t1, f4_2_t2, f4_2, "", by_name)}'
        f'</div>'
    )


# ─────────────────────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────────────────────

def _css():
    return """
<style>
@import url('https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;600;700&display=swap');

* { box-sizing: border-box; margin: 0; padding: 0; }

.brk-outer {
    background: #0E1117;
    padding: 14px 10px 20px;
    font-family: 'Barlow Condensed', 'Arial Narrow', Arial, sans-serif;
    overflow-x: auto;
    min-height: 100%;
}

.brk-title {
    text-align: center;
    font-size: 21px;
    font-weight: 700;
    letter-spacing: 6px;
    text-transform: uppercase;
    background: linear-gradient(90deg, #b89640, #f0d980, #b89640);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    padding-bottom: 3px;
}

.brk-owner {
    text-align: center;
    font-size: 11px;
    color: #3a4f6e;
    letter-spacing: 2px;
    margin-bottom: 12px;
}

.brk-main {
    display: flex;
    align-items: center;
    justify-content: center;
}

.brk-side {
    display: flex;
    flex-direction: column;
    gap: """ + str(REG_GAP) + """px;
}
</style>
"""


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def build_bracket_html(
    projections: pd.DataFrame,
    picks_row,
    bracket_owner: str = "",
    title: str = "🏀 MARCH MADNESS",
) -> str:
    """Return the full bracket as an HTML string."""
    by_name, by_rs = _build_lookups(projections)
    regions = list(dict.fromkeys(str(r) for r in projections["Region"]))[:4]

    left_html = (
        _region_html(projections, regions, 0, by_name, by_rs, picks_row, "left")
        + _region_html(projections, regions, 1, by_name, by_rs, picks_row, "left")
    )
    right_html = (
        _region_html(projections, regions, 2, by_name, by_rs, picks_row, "right")
        + _region_html(projections, regions, 3, by_name, by_rs, picks_row, "right")
    )
    center = _center_html(picks_row, by_name)

    return (
        _css()
        + f'<div class="brk-outer">'
        f'<div class="brk-title">{bracket_owner}</div>'
        f'<div class="brk-main">'
        f'<div class="brk-side">{left_html}</div>'
        f'{center}'
        f'<div class="brk-side">{right_html}</div>'
        f'</div>'
        f'</div>'
    )


def render_bracket(
    projections: pd.DataFrame,
    picks: pd.DataFrame = None,
    bracket_filter: str = None,
    title: str = "🏀 MARCH MADNESS",
    height: int = None,
):
    """
    Main Streamlit entry point.

    Parameters
    ----------
    projections    : 64-row team info DataFrame
    picks          : DataFrame with one row per bracket
    bracket_filter : pre-select a bracket by name (skips the selectbox)
    title          : display title above the bracket
    height         : iframe height in px (auto-calculated if None)
    """
    if picks is None or len(picks) == 0:
        st.warning("No picks data provided.")
        return

    if "Bracket" not in picks.columns:
        st.error("Picks DataFrame must have a 'Bracket' column.")
        return

    brackets = picks["Bracket"].dropna().unique().tolist()

    if bracket_filter and bracket_filter in brackets:
        selected = bracket_filter
    else:
        selected = st.selectbox("Select Bracket", brackets, key="brk_sel")

    row   = picks[picks["Bracket"] == selected].iloc[0]
    owner = str(selected)

    html = build_bracket_html(projections, row, bracket_owner=owner, title=title)

    # Auto-calculate height: 2 regions stacked + labels + gap + header + center
    if height is None:
        region_h = 16 * SLOT_H + LBL_H  # one region column height
        side_h   = 2 * region_h + REG_GAP + 14  # two regions + gap + region label
        height   = side_h + 140  # add some breathing room

    components.html(html, height=height, scrolling=True) #GD6